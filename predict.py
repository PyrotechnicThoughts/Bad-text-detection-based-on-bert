import os
from model import MyModel
from config import parsers
import torch
from transformers import BertTokenizer
import time
from fliter import remove
import pandas as pd
from openpyxl import load_workbook
import re

def load_model(device, model_path):
    myModel = MyModel().to(device)
    myModel.load_state_dict(torch.load(model_path))
    myModel.eval()
    return myModel


def process_text(text, bert_pred):
    tokenizer = BertTokenizer.from_pretrained(bert_pred)
    token_id = tokenizer.convert_tokens_to_ids(["[CLS]"] + tokenizer.tokenize(text))
    mask = [1] * len(token_id) + [0] * (args.max_len + 2 - len(token_id))
    token_ids = token_id + [0] * (args.max_len + 2 - len(token_id))
    token_ids = torch.tensor(token_ids).unsqueeze(0)
    mask = torch.tensor(mask).unsqueeze(0)
    x = torch.stack([token_ids, mask])
    return x


def text_class_name(pred):
    result = torch.argmax(pred, dim=1)
    result = result.cpu().numpy().tolist()
    classification = open(args.classification, "r", encoding="utf-8").read().split("\n")
    classification_dict = dict(zip(range(len(classification)), classification))
    return classification_dict[result[0]]

def convert_url_to_txt_filename(url):
    file_name = re.sub(r'[/:*?"<>|]', '_', url)  # 将不合法的字符替换为下划线
    return file_name

if __name__ == "__main__":
    start = time.time()
    args = parsers()
    device = "cuda:0" if torch.cuda.is_available() else "cpu"

    model = load_model(device, args.save_model_best)

    parent_folder = "/data/chengyn/xinan/Novel/TEXT/mobile/20240427"  # 文件夹名
    
    output_folder = "/data/chengyn/xinan/Novel/TEXT/分类结果/mobile/20240427"  # 文件夹名
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    
    bad_texts_folder = "/data/chengyn/xinan/Novel/TEXT/坏文本/mobile/20240427"  # 存放预测结果为 "bad" 的文本的文件夹名
    if not os.path.exists(bad_texts_folder):
        os.makedirs(bad_texts_folder)
        
    found_txt = False
    
    total_results = []
    for subdir, dirs, files in os.walk(parent_folder):  
        for file in files:
            if file.endswith(".txt"):
                badtxts=[]
                result_dict = remove(subdir)
                #print(result_dict)
                # 构建原始文件的完整路径
                file_path = os.path.join(subdir, file)
                print("读取txt",file_path)
                found_txt = True
                
                # 使用子文件夹名作为文件名前缀
                subfolder_name = os.path.basename(subdir)
                output_filename = os.path.join(output_folder, f"{subfolder_name}_classification.txt")
                bad_output_filename = os.path.join(bad_texts_folder, f"{subfolder_name}_bad_texts.txt")
                
                texts=list(set(result_dict['Pc-content.txt']))
  
                bad_text_count = 0
                total_text_count = len(texts)
                total_text_length = sum(len(text) for text in texts)
                print([subfolder_name, total_text_count])
        
                with open(output_filename, "w", encoding="utf-8") as file:  
                    for text in texts:
                        x = process_text(text, args.bert_pred)  # 假设 args.bert_pred 是处理参数
                        with torch.no_grad():
                            pred = model(x)  # 假设 model 是你的模型
                        classification = text_class_name(pred)
                        file.write(f"文本：{text}\t预测的类别为：{classification}\n")
                        
                        if classification == "bad_information":
                            badtxts.append(text)
                            bad_text_count += 1
                            
                    if bad_text_count>=5:
                        jud="yes"
                        with open(bad_output_filename, "w", encoding="utf-8") as bad_file:
                          for badtxt in badtxts:
                              bad_file.write(f"{badtxt}\n")
                    else:
                        jud="No"
                          
                        
                        
                    
                # 计算坏文本百分比
                if total_text_count != 0:
                    print(total_text_count != 0)
                    bad_text_percentage = round((bad_text_count / total_text_count) * 100, 2)
                    total_results.append([subfolder_name, total_text_count, bad_text_count, bad_text_percentage, total_text_length,jud])
                else:
                    total_results.append([subfolder_name, total_text_count, bad_text_count, 0, total_text_length,jud])
                  # 将结果添加到列表中
                

    
    
        # 如果在当前子文件夹中没有找到任何.txt文件
            if not found_txt:
            # 将文件夹名称和None值添加到结果列表中
                total_results.append((subfolder_name,0, 0, 0, 0,"No"))
                
    excel_file_path = "Novel.xlsx"
    wb = load_workbook(excel_file_path)
    ws = wb.active

    urls = [cell.value for cell in ws['A'][1:]]  # Assuming URLs are in column A
    results=[]

    # 遍历URL列表
    for website_url in urls:
        # 将URL转换为txt文件名
        txt_filename = convert_url_to_txt_filename(website_url)
        
        for result_list in total_results:
          key = result_list[0]
          if key==txt_filename:
            results.append([website_url] + result_list[1:])
        


    # 写入结果到 Excel 文件
    df = pd.DataFrame( results, columns=["url", "总text条数", "bad条数", "bad占比", "总长度",'是否为不良'])
    # 将 None 转换为字符串 "None"

    df.to_excel("分类结果统计表/Novel/mobile/分类统计结果_20240421.xlsx", index=False)
    
    end = time.time()
    print(f"耗时为：{end - start} s")
